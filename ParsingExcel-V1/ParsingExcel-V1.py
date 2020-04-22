import time
import uuid
import xlrd
from xlrd import open_workbook
from xlrd import dump
from xlrd.sheet import ctype_text
from openpyxl import Workbook
import json
import pika
import base64
import io
import re
from io import BytesIO
#import pandas as pd
import openpyxl
#import parser
import publisher
import datetime


credentials = pika.PlainCredentials(username='user', password='password')
connection = pika.BlockingConnection(
    pika.ConnectionParameters(host='54.144.236.23', port='5672', credentials=credentials))
channel = connection.channel()

channel.queue_declare(queue='hello', durable=False)
counter = 1




def callback(ch, method, properties, body):
    print(" [x] Received %r" % body)
    excel = body
    excel_dict = json.loads(excel)
    encoding_data = excel_dict["base64buffer"]
    decoded_data = base64.b64decode(encoding_data)
    file_extension = excel_dict["fileName"]
    userId = excel_dict["userId"]
    fileName = excel_dict["fileName"]
    curr_date = excel_dict["curr_date"]
    #message_format = {"userid": userId, "base64buffer": base64_message,
    #                  "date": curr_date, "fileName": filename}
    #message = json.dumps(message_format)
    print(userId)
    print(fileName)

    if "base64buffer" in excel_dict:
        if str(file_extension).split(".")[1] == 'xls':
            print('excel format : xls')
            try:
                wbb1 = Workbook()
                ws = wbb1.active
                #xl_workbook = open_workbook('c:\om\exampleMasterCard.xls') #           for read a file type
                xl_workbook = open_workbook(file_contents=decoded_data, on_demand=True)
                sheet_names = xl_workbook.sheet_names()
                #print(sheet_names)
                xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])

                xl_sheet = xl_workbook.sheet_by_index(0)
                #print('Sheet name: %s' % xl_sheet.name)
                row = xl_sheet.row(0)
                # print (row)
                #print('(Column #) type:value')
                for idx, cell_obj in enumerate(row):
                    cell_type_str = ctype_text.get(cell_obj.ctype, 'unknown type')
                    #print('(%s) %s %s' % (idx, cell_type_str, cell_obj.value))

                num_cols = xl_sheet.ncols  # Number of columns
                for row_idx in range(0, xl_sheet.nrows):  # Iterate through rows
                    #print('-' * 40)
                    #print('Row: %s' % row_idx)  # Print row number
                    for col_idx in range(0, num_cols):  # Iterate through columns
                        cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
                        #print('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
                        ws.cell(row=row_idx + 1, column=col_idx + 1).value = cell_obj.value
                try:
                    for row_idx in range(0, xl_sheet.nrows):  # Iterate through rows
                        #print('-' * 40)
                        #print('Row: %s' % row_idx)  # Print row number
                        for col_idx in range(0, 1):  # Iterate through columns
                            cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
                            #print('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
                            #print(cell_obj)
                            a = str(cell_obj).split(":")
                            #print(a[0])
                            c = 'xldate'
                            try:
                                if a[0] == 'xldate':
                                    #print('AAAAAAA')
                                    d = str(a[1]).split(".")
                                    e = int(d[0])
                                    #print(e)

                                    datetime_date = xlrd.xldate_as_datetime(e, 0)
                                    date_object = datetime_date.date()
                                    string_date = date_object.isoformat()
                                    date_time_originl = datetime.datetime.strptime(string_date, '%Y-%d-%m')
                                    date_time_convert = datetime.date.strftime(date_time_originl, "%d/%m/%y")
                                    #print(date_time_convert)
                                    ws.cell(row=row_idx + 1, column=col_idx + 1).value = date_time_convert

                            except:
                                continue
                except:
                    print("can't convert the first format ")
                try:
                    for row_idx in range(0, xl_sheet.nrows):  # Iterate through rows
                        #print('-' * 40)
                        #print('Row: %s' % row_idx)  # Print row number
                        for col_idx in range(1, 2):  # Iterate through columns
                            cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
                            #print('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
                            #print(cell_obj)
                            a = str(cell_obj).split(":")
                            c = 'xldate'
                            try:
                                if a[0] == 'xldate':
                                    d = str(a[1]).split(".")
                                    e = int(d[0])
                                    datetime_date = xlrd.xldate_as_datetime(e, 0)
                                    date_object = datetime_date.date()
                                    string_date = date_object.isoformat()
                                    try:
                                        date_time_originl = datetime.datetime.strptime(string_date, '%Y-%m-%d')
                                        date_time_convert = datetime.date.strftime(date_time_originl, "%d/%m/%y")

                                        ws.cell(row=row_idx + 1, column=col_idx + 1).value = date_time_convert
                                    except:
                                        ws.cell(row=row_idx + 1, column=col_idx + 1).value = date_time_convert
                            except:
                                continue
                except:
                    print("can't convert the second format ")

                f1 = 'c:\\OM\\output-xls-' + str(uuid.uuid4()) + '.xlsx'
                wbb1.save(filename=f1)

                workbook = openpyxl.load_workbook(f1)
                wb1 = workbook.active
                mr = wb1.max_row
                mc = wb1.max_column

                match = wb1.cell(row=1, column=1)
                match2 = wb1.cell(row=4, column=1)
                match3 = wb1.cell(row=3, column=1)

                if match2.value == None:
                    print("OTSAR HAHAYAL")
                    wb1.delete_cols(1)
                    wb1.delete_cols(3)
                    wbb = Workbook()
                    ws = wbb.active
                    x = ['תאריך עסקה', 'שם בית העסק', 'סכום עסקה', 'קטגוריה', 'מספר כרטיס']
                    counter = 1
                    for i in x:
                        ws.cell(row=1, column=counter, value=i)
                        counter += 1
                    card = wb1.cell(row=4, column=1)
                    s = str(card.value).split(":")[1].split(" ")
                    card_match = s[0]
                    for i in range(7, mr - 1):
                        for j in range(1, 4):
                            # reading cell value from source excel file
                            c = wb1.cell(row=i, column=j)

                            # writing the read value to destination excel file
                            ws.cell(row=i - 5, column=j).value = c.value
                            ws.cell(row=i - 5, column=5).value = card_match

                    for j in range(1, 4):
                        for i in range(1, mr):
                            if ws.cell(row=i, column=1).value is None:
                                ws.delete_rows(i)
                            else:
                                continue
                    f = 'c:\\OM\\output-otsar-hahayal-' + str(uuid.uuid4()) + '.xlsx'
                    wbb.save(filename=f)

                    with open(f, 'rb') as binary_file:
                        binary_file_data = binary_file.read()
                        base64_encoded_data = base64.b64encode(binary_file_data)
                        base64_message = base64_encoded_data.decode('utf-8')
                        print("Before publish to yogev")
                        message_format = {"userid": userId, "base64buffer": base64_message, "curr_date": curr_date,
                                           "fileName": fileName}
                        message = json.dumps(message_format)
                        publisher.sender(message)

                else:
                    if match.value == None:
                        print("ISRACARD")
                        wb1.delete_cols(3, 2)
                        sheet = wb1
                        alist = []
                        blist = []

                        wbb = Workbook()
                        ws = wbb.active
                        x = ['תאריך עסקה', 'שם בית העסק', 'סכום עסקה', 'קטגוריה', 'מספר כרטיס']
                        counter = 1
                        for i in x:
                            ws.cell(row=1, column=counter, value=i)
                            counter += 1
                        for i in range(4, mr):
                            b3 = sheet.cell(row=i, column=1)
                            s = str(b3.value).split(" ")
                            r = re.compile('[0-9]{4}')
                            newlist = list(filter(r.match, s))
                            list1 = [newlist, i]
                            if newlist != []:
                                alist.append(i)  # add row number for cars number to list
                                blist.append(', '.join(newlist))  # add cards number to list
                            else:
                                continue
                        alist.append(mr + 5)

                        for a in range(0, len(alist) - 1):
                            for i in range(alist[a] + 3, alist[a + 1] - 2):
                                for j in range(1, 4):
                                    # reading cell value from source excel file
                                    c = wb1.cell(row=i, column=j)
                                    # writing the read value to destination excel file
                                    ws.cell(row=i - 5, column=j).value = c.value
                                    ws.cell(row=i - 5, column=5).value = str(blist[a])

                        for j in range(1, 4):
                            for i in range(1, mr):
                                if ws.cell(row=i, column=1).value is None:
                                    ws.delete_rows(i)
                                else:
                                    continue
                        f = 'c:\\OM\\output-isracard-' + str(uuid.uuid4()) + '.xlsx'
                        wbb.save(filename=f)

                        with open(f, 'rb') as binary_file:
                            binary_file_data = binary_file.read()
                            base64_encoded_data = base64.b64encode(binary_file_data)
                            base64_message = base64_encoded_data.decode('utf-8')
                            print("Before publish to yogev")
                            message_format = {"userId": userId, "base64buffer": base64_message, "curr_date": curr_date,
                                              "fileName": fileName}
                            message = json.dumps(message_format)
                            print(message)
                            publisher.sender(message)
                            #print("send message to sender function")

                    elif match3.value!=None:
                        print("VISA")
                        wb1.delete_cols(3)
                        sheet = wb1
                        clist = []
                        b3 = sheet.cell(row=2, column=1)
                        s = str(b3.value).split(",")[1].split(" ")[3]
                        #print(s)
                        r = re.compile('[0-9]{4}')
                        # print(r)
                        newlist = list(s)
                        # rint(newlist)
                        list1 = [newlist, 2]
                        if newlist != []:
                            clist.append(', '.join(newlist))  # add cards number to list

                        wbb = Workbook()
                        ws = wbb.active
                        x = ['תאריך עסקה', 'שם בית העסק', 'סכום עסקה', 'קטגוריה', 'מספר כרטיס']
                        counter = 1
                        for i in x:
                            ws.cell(row=1, column=counter, value=i)
                            counter += 1

                        for i in range(4, mr):
                            for j in range(1, 4):
                                c = wb1.cell(row=i, column=j)
                                ws.cell(row=i - 2, column=j).value = c.value
                                ws.cell(row=i - 2, column=5).value = s

                        f = 'c:\\OM\\output-visa-' + str(uuid.uuid4()) + '.xlsx'
                        wbb.save(filename=f)

                        with open(f, 'rb') as binary_file:
                            binary_file_data = binary_file.read()
                            base64_encoded_data = base64.b64encode(binary_file_data)
                            base64_message = base64_encoded_data.decode('utf-8')
                            print("Before publish to yogev")

                            message_format = {"userId": userId, "base64buffer": base64_message, "curr_date": curr_date,
                                              "fileName": fileName}
                            message = json.dumps(message_format)

                            publisher.sender(message)
                    else:
                        print('cannot parsing file, Doesnt match any card format')
                        error_type = 'cannot parsing file, Doesnt match any card format'
                        message_error_format = {"userId": userId, "errtype": error_type, "curr_date": curr_date,
                                                "fileName": fileName}
                        message_err = json.dumps(message_error_format)
                        publisher.sender_err(message_err)

            except:
                print('cannot parsing to xlsx file, the xls file is corrupted')
                error_type = 'cannot parsing xls file, the file is corrupted'
                message_error_format = {"userId": userId, "errtype": error_type, "curr_date": curr_date,
                                  "fileName": fileName}
                message_err = json.dumps(message_error_format)
                publisher.sender_err(message_err)


        if str(file_extension).split(".")[1] == 'xlsx':
            print('excel format : xlsx')
            try:
                xls_filelike = io.BytesIO(decoded_data)
                workbook = openpyxl.load_workbook(xls_filelike)
                wbb2 = Workbook()
                ws = wbb2.active
                data = workbook
                # parser.read_file(data)
                f2 = 'c:\\OM\\output-xlsx-' + str(uuid.uuid4()) + '.xlsx'
                workbook.save(filename=f2)

                workbook = openpyxl.load_workbook(f2)
                wb1 = workbook.active
                mr = wb1.max_row
                mc = wb1.max_column

                match = wb1.cell(row=1, column=1)
                match2 = wb1.cell(row=4, column=1)
                match3 = wb1.cell(row=3, column=1)

                if match2.value == None:
                    print("OTSAR HAHAYAL")
                    wb1.delete_cols(1)
                    wb1.delete_cols(3)
                    card = wb1.cell(row=4, column=1)
                    s = str(card.value).split(":")[1].split(" ")
                    print(s[0])
                    card_match = s[0]

                    for i in range(7, mr - 1):
                        for j in range(1, 4):
                            # reading cell value from source excel file
                            c = wb1.cell(row=i, column=j)

                            # writing the read value to destination excel file
                            ws.cell(row=i - 5, column=j).value = c.value
                            ws.cell(row=i - 5, column=5).value = card_match

                    f = 'c:\\OM\\output-otsar-hahayal-' + str(uuid.uuid4()) + '.xlsx'
                    wbb.save(filename=f)

                    with open(f, 'rb') as binary_file:
                        binary_file_data = binary_file.read()
                        base64_encoded_data = base64.b64encode(binary_file_data)
                        base64_message = base64_encoded_data.decode('utf-8')
                        print("Before publish to yogev")

                        message_format = {"userId": userId, "base64buffer": base64_message, "curr_date": curr_date,
                                           "fileName": fileName}
                        message = json.dumps(message_format)

                        publisher.sender(message)
                else:
                    if match.value==None:
                        print("ISRACARD")
                        wb1.delete_cols(3, 2)
                        sheet = wb1
                        alist = []
                        blist = []

                        wbb = Workbook()
                        ws = wbb.active
                        x = ['תאריך עסקה', 'שם בית העסק', 'סכום עסקה', 'קטגוריה', 'מספר כרטיס']
                        counter = 1
                        for i in x:
                            ws.cell(row=1, column=counter, value=i)
                            counter += 1
                        for i in range(4, mr):
                            b3 = sheet.cell(row=i, column=1)
                            s = str(b3.value).split(" ")
                            r = re.compile('[0-9]{4}')
                            newlist = list(filter(r.match, s))
                            list1 = [newlist, i]
                            if newlist != []:
                                alist.append(i)  # add row number for cars number to list
                                blist.append(', '.join(newlist))  # add cards number to list
                            else:
                                continue
                        alist.append(mr + 5)

                        for a in range(0, len(alist) - 1):
                            for i in range(alist[a] + 3, alist[a + 1] - 2):
                                for j in range(1, 4):
                                    # reading cell value from source excel file
                                    c = wb1.cell(row=i, column=j)
                                    # writing the read value to destination excel file
                                    ws.cell(row=i - 5, column=j).value = c.value
                                    ws.cell(row=i - 5, column=5).value = str(blist[a])

                        for j in range(1, 4):
                            for i in range(1, mr):
                                if ws.cell(row=i, column=1).value is None:
                                    ws.delete_rows(i)
                                else:
                                    continue
                        f = 'c:\\OM\\output-isracard' + str(uuid.uuid4())+'.xlsx'
                        wbb.save(filename=f)

                        #channel.queue_declare(queue='Yogevhello', durable=True)
                        with open(f, 'rb') as binary_file:
                            # print(binary_file)
                            binary_file_data = binary_file.read()
                            base64_encoded_data = base64.b64encode(binary_file_data)
                            base64_message = base64_encoded_data.decode('utf-8')
                            print("Before publish to yogev")
                            message_format = {"userId": userId, "base64buffer": base64_message, "curr_date": curr_date,
                                               "fileName": fileName}
                            message = json.dumps(message_format)
                            publisher.sender(message)

                    if match3.value==None:
                        #match3.value==None:
                        print("VISA")
                        wb1.delete_cols(3)
                        sheet = wb1
                        clist = []
                        b3 = sheet.cell(row=2, column=1)
                        s = str(b3.value).split(",")[1].split(" ")[3]
                        print(s)
                        r = re.compile('[0-9]{4}')
                        #print(r)
                        newlist = list(s)
                        #rint(newlist)
                        list1 = [newlist, 2]
                        if newlist != []:
                            clist.append(', '.join(newlist))  # add cards number to list


                        wbb = Workbook()
                        ws = wbb.active
                        x = ['תאריך עסקה', 'שם בית העסק', 'סכום עסקה', 'קטגוריה', 'מספר כרטיס']
                        counter = 1
                        for i in x:
                            ws.cell(row=1, column=counter, value=i)
                            counter+=1

                        for i in range(4,mr):
                            for j in range(1,4):
                                c = wb1.cell(row=i, column=j)
                                ws.cell(row=i - 2, column=j).value = c.value
                                ws.cell(row=i - 2, column=5).value = s

                        f = 'c:\\OM\\output-visa-' + str(uuid.uuid4()) + '.xlsx'
                        wbb.save(filename=f)

                        with open(f, 'rb') as binary_file:
                            binary_file_data = binary_file.read()
                            base64_encoded_data = base64.b64encode(binary_file_data)
                            base64_message = base64_encoded_data.decode('utf-8')
                            print("Before publish to yogev")

                            message_format = {"userId": userId, "base64buffer": base64_message, "curr_date": curr_date,
                                               "fileName": fileName}
                            message = json.dumps(message_format)

                            publisher.sender(message)
                    else:
                        print('cannot parsing file, Doesnt match any card format')
                        error_type = 'cannot parsing file, Doesnt match any card format'
                        message_error_format = {"userId": userId, "errtype": error_type, "curr_date": curr_date,
                                                "fileName": fileName}
                        message_err = json.dumps(message_error_format)
                        publisher.sender_err(message_err)
            except:
                print('cannot parsing xlsx file, the file is corrupted')
                error_type = 'cannot parsing xls file, the file is corrupted'
                message_error_format = {"userId": userId, "errtype": error_type, "curr_date": curr_date,
                                  "fileName": fileName}
                message_err = json.dumps(message_error_format)
                publisher.sender_err(message_err)
    else:
        print('Cannot parsing json from rabbit')
        error_type = 'cannot parsing Json Message, the Json Message is corrupted'
        message_error_format = {"userId": userId, "errtype": error_type, "curr_date": curr_date,
                                "fileName": fileName}
        message_err = json.dumps(message_error_format)
        publisher.sender_err(message_err)


# channel.basic_consume(
#     queue='userDataToProcess', on_message_callback=callback, auto_ack=True)
#
# print(' [*] Waiting for messages. To exit press CTRL+C')
#
# channel.start_consuming()


while(True):
    try:
        print("Connecting...")
        channel.basic_consume(
            queue='q_eldar', on_message_callback=callback, auto_ack=True)
        print(' [*] Waiting for messages. To exit press CTRL+C')
        try:
            channel.start_consuming()
            print("1 - start consuming")
        except KeyboardInterrupt:
            channel.stop_consuming()
            connection.close()
            time.sleep(5)
            print("2 - stop consuming and close connction")
            continue
    except pika.exceptions.ConnectionClosedByBroker:
        # Uncomment this to make the example not attempt recovery
        # from server-initiated connection closure, including
        # when the node is stopped cleanly
        #
        channel.start_consuming()
        print("3 - start consuming after exeption 1 closed by broker")
        continue
    # Do not recover on channel errors
    except pika.exceptions.AMQPChannelError as err:
        print("Caught a channel error: {}, stopping...".format(err))
        channel.stop_consuming()
        connection.close()
        time.sleep(5)
        print("4 - stop consuming and close connction")
        channel.start_consuming()
        print("4 - start consuming after connection closed")
        continue
    #Recover on all other connection errors
    except pika.exceptions.AMQPConnectionError:
        print("Connection was closed, retrying...")
        print("5 - retry connect")
        time.sleep(5)
        # channel.start_consuming()
        continue