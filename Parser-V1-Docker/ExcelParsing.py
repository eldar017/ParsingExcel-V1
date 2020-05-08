import time
import uuid
import xlrd
from xlrd import open_workbook
from xlrd import dump
from xlrd.sheet import ctype_text
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import pika
import base64
import io
import re
from io import BytesIO
import openpyxl
import publisher
import datetime
import Logger

class Main:

    def file_type(self):
        self._raw1 = self._wb1.cell(row=4, column=2)
        self._raw2 = self._wb1.cell(row=3, column=1)
        self._raw3 = self._wb1.cell(row=1, column=1)
        self._raw4 = self._wb1.cell(row=21, column=1)
        self._raw4_2 = self._wb1.cell(row=22, column=1)
        self._raw5 = self._wb1.cell(row=6, column=6)
        try:
            if self._raw1.value=='מועד חיוב':
                print("ISRACARD")
                Logger.logger.info('ISRACARD')
                self.isracard()
            elif self._raw2.value=='תאריך העסקה':
                print("VISA")
                Logger.logger.info('VISA')
                self.visa()
            elif self._raw3.value=='כרטיסי אשראי':
                print("DISCOUNT")
                Logger.logger.info('DISCOUNT')
                self.discount()
            elif self._raw4.value=='פירוט עבור הכרטיסים בארץ':
                print("POALIM")
                Logger.logger.info('POALIM')
                self.poalim()
            elif self._raw4_2.value=='פירוט עבור הכרטיסים בארץ':
                print("POALIM")
                Logger.logger.info('POALIM')
                self.poalim()
            elif self._raw5.value == 'פירוט':
                print("OTSAR HAHAYAL")
                Logger.logger.info('OTSAR HAHAYAL')
                self.otshar_hahayal()
            else:
                print("cannot convert, the excel file isNot compatible with any template")
                Logger.logger.debug('cannot convert, the excel file isNot compatible with any template')
        except:
            print("except - cannot convert, the excel file isNot compatible with any template")
            Logger.logger.debug('except - cannot convert, the excel file isNot compatible with any template')
            self._error_type = "except - cannot convert, the excel file isNot compatible with any template"
            self.build_err_message()

    def build_message(self):
        try:
            with open(self._f, 'rb') as binary_file:
                # print(binary_file)
                self._binary_file_data = binary_file.read()
                self._base64_encoded_data = base64.b64encode(self._binary_file_data)
                self._base64_message = self._base64_encoded_data.decode('utf-8')
                print("Before publish to yogev")
                self._message_format = {"userId": self._userId, "base64buffer": self._base64_message, "curr_date": self._curr_date,
                                  "fileName": self._fileName}
                self._message = json.dumps(self._message_format)
                publisher.sender(self._message)
        except:
            print("cannot build_message")
            Logger.logger.debug('cannot build_message')
            self._error_type = "cannot build_message"
            self.build_err_message()

    def build_err_message(self):
        try:
            self._message_err_format = {"userId": self._userId, "errtype": self._error_type, "curr_date": self._curr_date,
                              "fileName": self._fileName}
            self._message_err = json.dumps(self._message_err_format)
            publisher.sender_err(self._message_err)
        except:
            print("cannot build_message_err")
            Logger.logger.debug('cannot build_message_err')
            self._error_type = "cannot build_message_err"
            self.build_err_message()

    def file_extension(self, body):
        try:
            self._path = '/opt/ExcelBackup/'
            self._excel = body
            self._excel_dict = json.loads(self._excel)
            self._encoding_data = self._excel_dict["base64buffer"]
            self._decoded_data = base64.b64decode(self._encoding_data)
            self._file_extension = self._excel_dict["fileName"]
            self._userId = self._excel_dict["userId"]
            self._fileName = self._excel_dict["fileName"]
            self._curr_date = self._excel_dict["curr_date"]
            Logger.logger.info(self._userId)
            print(self._userId)
            Logger.logger.info(self._fileName)
            print(self._fileName)
            if "base64buffer" in self._excel_dict:
                if str(self._file_extension).split(".")[1] == 'xls':
                    print('excel format : xls')
                    Logger.logger.info('excel format : xls')
                    self.convert_xls()
                elif str(self._file_extension).split(".")[1] == 'xlsx':
                    print('excel format : xlsx')
                    Logger.logger.info('excel format : xlsx')
                    self.read_xlsx()
        except:
            print("problem with parsing message from rabbit - file_extention function")
            Logger.logger.debug('problem with parsing message from rabbit - file_extention function')
            self._error_type = "problem with parsing message from rabbit - file_extention function"
            self.build_err_message()


    def read_file(self):
        #self._fileName = "C:\OM\exampleDiscount.xlsx"
        #self._fileName = "C:\OM\exampleVisaCard.xlsx"
        #self._fileName = "C:\OM\exampleMasterCard.xlsx"
        #self._fileName = "C:\OM\Export_9_03_2020.xlsx"
        #self._fileName = "C:\OM\examleOtsharHahayal.xls"
        #self._fileName = "C:\OM\excelNewBank.xlsx"

        self._wb = load_workbook(self._fileName)
        #self._workbook = workbook
        self._wb1 = self._wb.active
        self._mr = self._wb1.max_row
        self._mc = self._wb1.max_column
        self.file_type()

    def otshar_hahayal(self):
        try:
            self._wb1.delete_cols(1)
            self._wb1.delete_cols(3)
            self.output_file()
            self._card = self._wb1.cell(row=4, column=1)
            self._s = str(self._card.value).split(":")[1].split(" ")
            self._card_match = self._s[0]
            for i in range(7, self._mr - 1):
                for j in range(1, 4):
                    self._c = self._wb1.cell(row=i, column=j)
                        # writing the read value to destination excel file
                    self._ws.cell(row=i - 5, column=j).value = self._c.value
                    self._ws.cell(row=i - 5, column=5).value = self._card_match

            for j in range(1, 4):
                for i in range(1, self._mr):
                    if self._ws.cell(row=i, column=1).value is None:
                        self._ws.delete_rows(i)
                    else:
                        continue
            self._f = self._path + 'output-otsar-hahayal-' + str(uuid.uuid4()) + '.xlsx'
            self._wb.save(self._f)
            self.build_message()
        except:
            print("cannot parsing otsar-hahayal template ")
            Logger.logger.debug('cannot parsing otsar-hahayal template')
            self._error_type = "cannot parsing otsar-hahayal template"
            self.build_err_message()

    def poalim(self):
        try:
            self._wb1.delete_cols(2)
            self._wb1.delete_cols(6,13)
            self.output_file()
            self._clist = []
            self._dlist = []

            for i in range(1,self._mr-2):
                self._b4 = self._wb1.cell(row=i, column=1).value
                if  self._b4 =='פירוט עבור הכרטיסים בארץ':
                    self._clist.append(i)
                elif self._b4 =="פירוט עבור הכרטיסים בחו''ל בדולר":
                    self._dlist.append(i)
                else:
                    continue

            for i in range(self._clist[0]+3, self._dlist[0]-2):
                #print( self._wb1.cell(row=i,column=3).value)
                for j in range(1, 6):
                    # reading cell value from source excel file
                    self._c = self._wb1.cell(row=i, column=j)
                    self._ca = self._wb1.cell(row=i, column=1)
                    # writing the read value to destination excel file
                    if j == 1:
                        self._ws.cell(row=i - 2, column=5).value = self._ca.value
                    elif j == 2:
                        #convert time forame from dd-mm-yyy to dd/mm/yyyy
                        self._date_time = self._c.value
                        self._d = self._date_time.strftime("%d/%m/%Y, %H:%M:%S")
                        self._dd = str(self._d).split(",")[0]
                        self._ws.cell(row=i - 2, column=1).value = self._dd
                    elif j == 3:
                        self._ws.cell(row=i - 2, column=2).value = self._c.value
                    elif j == 4:
                        self._ws.cell(row=i - 2, column=3).value = self._c.value
                    else:
                        continue

            self._ws.delete_rows(self._dlist[0]-2,self._mr)
            self._ws.delete_rows(2, self._clist[0]-1)
            #self._ws.delete_cols(6,9)
            self._f = self._path + 'output-poalim-' + str(uuid.uuid4()) + '.xlsx'
            self._wb.save(self._f)
            self.build_message()
        except:
            print("cannot parsing poalim template")
            Logger.logger.debug('cannot parsing poalim template')
            self._error_type = "cannot parsing poalim template"
            self.build_err_message()

    def discount(self):
        try:
            self._wb1.delete_cols(4,5)
            self.output_file()
            for i in range(18,self._mr-2):
                #print( self._wb1.cell(row=i,column=4).value)
                for j in range(2, 10):
                    # reading cell value from source excel file
                    self._c = self._wb1.cell(row=i, column=j)
                    self._ca = self._wb1.cell(row=i, column=1)
                    self._s = str(self._ca.value).split(" ")[1]
                    # writing the read value to destination excel file
                    if j == 3:
                        self._ws.cell(row=i - 2, column=1).value = self._c.value
                    elif j == 4:
                        self._ws.cell(row=i - 2, column=3).value = self._c.value
                    else:
                        self._ws.cell(row=i - 2, column=j).value = self._c.value
                    self._ws.cell(row=i - 2, column=5).value = self._s
            self._ws.delete_rows(2,14)
            self._ws.delete_cols(6,9)
            self._f = self._path + 'output-discount-' + str(uuid.uuid4()) + '.xlsx'
            self._wb.save(self._f)
            self.build_message()
        except:
            print("cannot parsing discount template")
            Logger.logger.debug('cannot parsing discount template')
            self._error_type = "cannot parsing discount template"
            self.build_err_message()

    def visa(self):
        try:
            self._wb1.delete_cols(3)
            self._clist = []
            self._b3 = self._wb1.cell(row=2, column=1)
            self._s = str(self._b3.value).split(",")[1].split(" ")[3]
            self.output_file()
            for i in range(4,self._mr):
                for j in range(1, 4):
                    # reading cell value from source excel file
                    self._c = self._wb1.cell(row=i, column=j)

                    # writing the read value to destination excel file
                    self._ws.cell(row=i - 2, column=j).value = self._c.value
                    self._ws.cell(row=i - 2, column=5).value = self._s

            self._f = self._path + 'output-visa-' + str(uuid.uuid4()) + '.xlsx'
            self._wb.save(self._f)
            self.build_message()
        except:
            print("cannot parsing visa template")
            Logger.logger.debug('cannot parsing visa template')
            self._error_type = "cannot parsing visa template"
            self.build_err_message()


    def isracard(self):
        try:
            self._sheet=self._wb1
            self._alist=[]
            self._blist=[]
            self._clist=[]
            self._dlist = []
            self.output_file()
            for i in range(4, self._mr):
                self._b3 = self._sheet.cell(row=i, column=1)
                self._b4 = self._sheet.cell(row=i, column=2)
                self._b5 = self._sheet.cell(row=i, column=3)
                self._s = str(self._b3.value).split(" ")
                self._r = re.compile('[0-9]{4}')
                self._newlist = list(filter(self._r.match, self._s))
                self._list1 = [self._newlist, i]
                if self._newlist != []:
                    self._alist.append(i)  # add row number for cars number to list
                    self._blist.append(', '.join(self._newlist))  # add cars number to list
                elif self._b4.value =='תאריך חיוב':
                    self._clist.append(i)
                elif self._b5.value =='TOTAL FOR DATE':
                    self._dlist.append(i)
                else:
                    continue

            self._wb1.delete_cols(3, 2)
            self._alist.append(self._mr + 5)
            # insert the value from the original excel
            for a in range(0, len(self._alist) - 1):
                for i in range(self._alist[a] + 3, self._alist[a + 1] - 2):
                    for j in range(1, 4):
                        # reading cell value from source excel file
                        self._c = self._wb1.cell(row=i, column=j)
                        # writing the read value to destination excel file
                        self._ws.cell(row=i - 5, column=j).value = self._c.value
                        self._ws.cell(row=i - 5, column=5).value = str(self._blist[a])

            for j in range(1, 4):
                for i in range(1, self._mr):
                    if self._ws.cell(row=i, column=1).value == "":
                        self._ws.delete_rows(i)
                    elif self._ws.cell(row=i, column=1).value == None:
                        self._ws.delete_rows(i)
                    elif self._ws.cell(row=i, column=3).value == '₪':
                        self._ws.delete_rows(i)
                    else:
                        continue
            count =1
            if len(self._clist) != '0':
                for i in range(0, len(self._clist)):
                    self._ws.delete_rows(self._clist[i] - 7)
                    self._ws.delete_rows(self._clist[i] - 8)
            else:
                count+=1
            self._f = self._path + 'output-isracard-' + str(uuid.uuid4()) + '.xlsx'
            self._wb.save(filename=self._f)
            self.build_message()

        except:
            print("cannot parsing isracard template")
            Logger.logger.debug('cannot parsing isracard template')
            self._error_type = "cannot parsing isracard template"
            self.build_err_message()


    def convert_xls(self):
         try:
            self._wbb1 = Workbook()
            self._ws = self._wbb1.active
            # xl_workbook = open_workbook('c:\om\exampleMasterCard.xls') #           for read a file type
            self._xl_workbook = open_workbook(file_contents=self._decoded_data, on_demand=True)
            self._sheet_names = self._xl_workbook.sheet_names()
            # print(sheet_names)
            self._l_sheet = self._xl_workbook.sheet_by_name(self._sheet_names[0])

            self._xl_sheet = self._xl_workbook.sheet_by_index(0)
            # print('Sheet name: %s' % xl_sheet.name)
            self._row = self._xl_sheet.row(0)
            for idx, cell_obj in enumerate(self._row):
                self._cell_type_str = ctype_text.get(cell_obj.ctype, 'unknown type')
                # print('(%s) %s %s' % (idx, cell_type_str, cell_obj.value))

            self._num_cols = self._xl_sheet.ncols  # Number of columns
            for row_idx in range(0, self._xl_sheet.nrows):  # Iterate through rows
                # print('-' * 40)
                # print('Row: %s' % row_idx)  # Print row number
                for col_idx in range(0, self._num_cols):  # Iterate through columns
                    self._cell_obj = self._xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
                    # print('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
                    self._ws.cell(row=row_idx + 1, column=col_idx + 1).value = self._cell_obj.value
            try:
                for row_idx in range(0, self._xl_sheet.nrows):  # Iterate through rows
                    #print('-' * 40)
                    # print('Row: %s' % row_idx)  # Print row number
                    for col_idx in range(1, 2):  # Iterate through columns
                        self._cell_obj = self._xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
                        #print('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))
                        self._a = str(self._cell_obj).split(":")
                        self._c = 'xldate'
                        try:
                            if self._a[0] == 'xldate':
                                self._d = str(self._a[1]).split(".")
                                self._e = int(self._d[0])

                                self._datetime_date = xlrd.xldate_as_datetime(self._e, 0)
                                self._date_object = self._datetime_date.date()
                                self._string_date = self._date_object.isoformat()

                                try:
                                    self._date_time_originl = datetime.datetime.strptime(self._string_date, "%Y-%d-%m")
                                    self._date_time_convert = self._date_time_originl.strftime("%d/%m/%Y")

                                    self._ws.cell(row=row_idx + 1, column=col_idx + 1).value = self._date_time_convert
                                except:
                                    self._ws.cell(row=row_idx + 1, column=col_idx + 1).value = self._date_time_convert

                        except:
                            continue
            except:
                print("can't convert the first format ")
                Logger.logger.debug('cannot convert the first format')
                self._error_type = "can't convert the first format"
                self.build_err_message()
            try:
                for row_idx in range(0, self._xl_sheet.nrows):  # Iterate through rows
                    #print('-' * 40)
                    # print('Row: %s' % row_idx)  # Print row number
                    for col_idx in range(0, 1):  # Iterate through columns
                        self._cell_obj = self._xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col

                        self._a = str(self._cell_obj).split(":")
                        self._c = 'xldate'
                        try:
                            if self._a[0] == 'xldate':
                                self._d = str(self._a[1]).split(".")
                                self._e = int(self._d[0])
                                self._datetime_date = xlrd.xldate_as_datetime(self._e, 0)
                                self._date_object = self._datetime_date.date()
                                self._string_date = self._date_object.isoformat()
                                try:
                                    self._date_time_originl = datetime.datetime.strptime(self._string_date, "%Y-%d-%m")
                                    self._date_time_convert = self._date_time_originl.strftime("%d/%m/%Y")

                                    self._ws.cell(row=row_idx + 1, column=col_idx + 1).value = self._date_time_convert
                                except:
                                    self._ws.cell(row=row_idx + 1, column=col_idx + 1).value = self._date_time_convert
                        except:
                            continue
            except:
                print("can't convert the second format ")
                Logger.logger.debug('cannot convert the first format')
                self._error_type = "can't convert the second format"
                self.build_err_message()

            self._f1 = self._path + 'output-xls-' + str(uuid.uuid4()) + '.xlsx'
            self._wbb1.save(filename=self._f1)

            self._workbook = openpyxl.load_workbook(self._f1)
            self._wb1 = self._workbook.active
            self._mr = self._wb1.max_row
            self._mc = self._wb1.max_column
            #print("!@!@!@!@")
            self.file_type()
         except:
             print("cannot convert xls file")
             Logger.logger.debug('cannot convert xls file')
             self._error_type = "cannot convert xls file"
             self.build_err_message()

    def read_xlsx(self):
        try:
            self._xls_filelike = io.BytesIO(self._decoded_data)
            self._workbook = openpyxl.load_workbook(self._xls_filelike)
            self._wbb2 = Workbook()
            self._ws = self._wbb2.active
            self._data = self._workbook
            # parser.read_file(data)
            self._f1 = self._path + str(uuid.uuid4()) + '.xlsx'
            self._workbook.save(filename=self._f1)

            self._workbook = openpyxl.load_workbook(self._f1)
            self._wb1 = self._workbook.active
            self._mr = self._wb1.max_row
            self._mc = self._wb1.max_column
            self.file_type()
        except:
            print("cannot read excel file from binary")
            Logger.logger.debug('cannot read excel file from binary')
            self._error_type = "cannot read excel file from binary"
            self.build_err_message()

    def output_file(self):
        try:
            self._wb = Workbook()
            self._ws = self._wb.active
            self._x = ['תאריך עסקה', 'שם בית העסק', 'סכום עסקה', 'קטגוריה', 'מספר כרטיס']
            self._counter = 1
            for i in self._x:
                self._ws.cell(row=1, column=self._counter, value=i)
                self._counter += 1
        except:
            print("cannot make a new file for output")
            Logger.logger.debug('cannot make a new file for output')
            self._error_type = "cannot make a new file for output"
            self.build_err_message()

Object = Main()
#Object.read_file()
#Object.callback()